import asyncio
import re
import subprocess
from datetime import datetime
from zoneinfo import ZoneInfo

from discord import File
from discord.ext import commands
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from wordle.settings import get_settings

POINT_MAP = {
    "1": 10,
    "2": 5,
    "3": 4,
    "4": 3,
    "5": 2,
    "6": 1,
    "X": -1,
}


class EventsCog(commands.Cog):
    """This is a cog for event listeners.

    Evnt listeners are executed automatically on a specific event.
    They can be used to handle errors, guild changes and more.
    """

    def __init__(self, bot: commands.Bot, month: str):
        self.bot = bot
        self.month = datetime.strptime(month, "%Y-%m")

        self.workbook_filename = "wordle_stats.xlsx"
        self.workbook_image = "wordle_stats.png"

    @commands.Cog.listener()
    async def on_ready(self) -> None:
        logger.info("Bot is ready.")

        asyncio.create_task(self.parse_wordle_stats())

    async def parse_wordle_stats(self):
        """Read messages from the wordle bot and generate statistics."""

        await asyncio.sleep(0)

        user_data = await self.get_user_data()
        if not user_data:
            await self.bot.close()
            return

        await self.generate_report(user_data)
        await self.genarate_image()

        attachments = [File(self.workbook_image), File(self.workbook_filename)]
        await self._channel.send(
            "Wordle statistics for this month :)",
            files=attachments,
        )

        await self.bot.close()

    async def get_user_data(self) -> dict[str, dict[str, int]]:
        """Parse daily statistics and calculate total user score."""

        settings = get_settings()

        guild = self.bot.get_guild(settings.guild_id)
        if guild is None:
            logger.error("Guild not found. Is the bot in the server?")
            return {}

        self._channel = guild.get_channel(settings.channel_id)
        if self._channel is None:
            logger.error("Channel not found.")
            return {}

        user_data = {}

        points_pattern = r".*([0-6X])/[0-6]:"

        # Get start and end datetimes for the messages
        start = datetime(
            self.month.year,
            self.month.month,
            1,
            tzinfo=ZoneInfo(settings.timezone),
        )

        if self.month.month == 12:
            end = datetime(
                self.month.year + 1,
                1,
                1,
                tzinfo=ZoneInfo(settings.timezone),
            )
        else:
            end = datetime(
                self.month.year,
                self.month.month + 1,
                1,
                tzinfo=ZoneInfo(settings.timezone),
            )

        channel_messages = self._channel.history(
            limit=None,
            after=start,
            before=end,
            oldest_first=True,
        )

        async for message in channel_messages:
            if message.author.id != settings.app_id:
                continue

            message_date = message.created_at.astimezone(ZoneInfo(settings.timezone))
            date = ""
            if message_date is not None:
                date = message_date.strftime("%d %a")

            if date not in user_data:
                user_data[date] = {}

            message_mentions = {
                str(member.id): member.display_name for member in message.mentions
            }

            content = message.content.strip()
            if "results:" not in content:
                continue

            # Split the content into lines and iterate over it
            lines = content.split("\n")

            for line in lines:
                points = re.findall(points_pattern, line)
                if points == []:
                    continue

                score = POINT_MAP[points[0]]

                for user_id, display_name in message_mentions.items():
                    if user_id in line or display_name in line:
                        if display_name not in user_data[date]:
                            user_data[date][display_name] = 0

                        user_data[date][display_name] += score

        logger.info(user_data)

        return user_data

    async def generate_report(self, user_data: dict[str, dict[str, int]]) -> None:
        """Create excel sheet with user data."""

        wb = Workbook()
        ws = wb.active
        if ws is None:
            logger.error("Could not get active sheet")
            return

        # Configure sheet style to be rendered by libreoffice
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_area = "A1:AF34"
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25)

        ws.title = "Wordle Statistics"

        header_fill = PatternFill("solid", fgColor="8A8A8A")
        data_fill = PatternFill("solid", fgColor="BABABA")
        extra_headers_fill = PatternFill("solid", fgColor="75BB75")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        no_border = Border(
            left=Side(style="none"),
            right=Side(style="none"),
            top=Side(style="none"),
            bottom=Side(style="none"),
        )

        # Placement medal colors
        gold_fill = PatternFill("solid", fgColor="FFD700")
        silver_fill = PatternFill("solid", fgColor="C0C0C0")
        bronze_fill = PatternFill("solid", fgColor="CD7F32")

        usernames: set[str] = set()
        for day_data in user_data.values():
            usernames.update(day_data.keys())

        users = sorted(usernames)
        days = list(user_data.keys())

        # Format column for usernames
        ws["A1"] = "Players"
        ws["A1"].fill = header_fill
        ws["A1"].border = no_border

        totals: dict[str, int] = {}

        for row_index, username in enumerate(users, start=2):
            username_cell = ws[f"A{row_index}"]
            username_cell.value = username
            username_cell.border = no_border
            username_cell.fill = header_fill

            row_total = 0
            for col_index, day in enumerate(days, start=2):
                day_score = user_data.get(day, {}).get(username, 0)
                header = ws.cell(row=1, column=col_index)
                header.value = day
                header.fill = header_fill
                header.border = no_border
                header.alignment = align_right

                cell = ws.cell(row=row_index, column=col_index)
                cell.value = day_score
                cell.fill = data_fill
                cell.border = no_border
                row_total += day_score

            totals[username] = row_total

        last_main_row = len(users) + 1
        buffer_row = last_main_row + 1
        totals_header_row = buffer_row + 1

        # Merge the A+B columns and write the "TOTAL SCORE" header
        ws.merge_cells(
            start_row=totals_header_row,
            start_column=1,
            end_row=totals_header_row,
            end_column=2,
        )
        total_title_cell = ws.cell(row=totals_header_row, column=1)
        total_title_cell.value = "TOTAL SCORE"
        total_title_cell.fill = extra_headers_fill
        total_title_cell.border = no_border
        total_title_cell.alignment = align_center

        # Write sub-headers A: "Players" and B: "Total Points"
        totals_subheaders_row = totals_header_row + 1

        players_header = ws.cell(row=totals_subheaders_row, column=1)
        players_header.value = "Players"
        players_header.fill = header_fill
        players_header.border = no_border

        points_header = ws.cell(row=totals_subheaders_row, column=2)
        points_header.value = "Total Points"
        points_header.fill = header_fill
        points_header.border = no_border

        # Write data to the totals table
        totals_data_start_row = totals_subheaders_row + 1

        for offset, username in enumerate(users):
            row = totals_data_start_row + offset

            name_cell = ws.cell(row=row, column=1)
            name_cell.value = username
            name_cell.fill = header_fill
            name_cell.border = no_border

            total_cell = ws.cell(row=row, column=2)
            total_cell.value = totals.get(username, 0)
            total_cell.fill = data_fill
            total_cell.border = no_border

        # Create a leaderboard table with each player's placement.
        # First sort the totals
        sorted_totals = sorted(totals.items(), key=lambda k: k[1], reverse=True)

        leaderboard_start_col = 4
        leaderboard_header_row = totals_header_row

        ws.merge_cells(
            start_row=leaderboard_header_row,
            start_column=leaderboard_start_col,
            end_row=leaderboard_header_row,
            end_column=leaderboard_start_col + 1,
        )
        leaderboard_title_cell = ws.cell(
            row=leaderboard_header_row,
            column=leaderboard_start_col,
        )
        leaderboard_title_cell.value = "WORDLE LEADERBOARD"
        leaderboard_title_cell.fill = extra_headers_fill
        leaderboard_title_cell.border = no_border
        leaderboard_title_cell.alignment = align_center

        leaderboard_subheader_row = leaderboard_header_row + 1

        lb_players_header = ws.cell(
            row=leaderboard_subheader_row,
            column=leaderboard_start_col,
        )
        lb_players_header.value = "Players"
        lb_players_header.fill = header_fill
        lb_players_header.border = no_border

        lb_placement_header = ws.cell(
            row=leaderboard_subheader_row,
            column=leaderboard_start_col + 1,
        )
        lb_placement_header.value = "Placement"
        lb_placement_header.fill = header_fill
        lb_placement_header.border = no_border

        # Leaderboard data rows
        leaderboard_data_start_row = leaderboard_subheader_row + 1

        for rank, (username, total) in enumerate(sorted_totals, start=1):
            row = leaderboard_data_start_row + (rank - 1)

            player_cell = ws.cell(row=row, column=leaderboard_start_col)
            player_cell.value = username
            player_cell.border = no_border
            player_cell.fill = header_fill

            placement_cell = ws.cell(row=row, column=leaderboard_start_col + 1)
            placement_cell.value = f"#{rank}"
            placement_cell.border = no_border
            placement_cell.fill = data_fill

            if rank == 1:
                player_cell.fill = gold_fill
                placement_cell.fill = gold_fill
            elif rank == 2:
                player_cell.fill = silver_fill
                placement_cell.fill = silver_fill
            elif rank == 3:
                player_cell.fill = bronze_fill
                placement_cell.fill = bronze_fill

        # Auto-fit column widths for usernames
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = max_length + 2

        ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        wb.save(self.workbook_filename)
        logger.info(f"Saved excel report to '{self.workbook_filename}'")

    async def genarate_image(self) -> None:
        """Convert the excel sheet to a PDF and then to a PNG."""

        # Use libreoffice to generate a PDF from the workbook
        libreoffice_command = [
            "libreoffice",
            "--headless",
            "--convert-to",
            "png",
            self.workbook_filename,
        ]
        subprocess.run(libreoffice_command, check=True)
